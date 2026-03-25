"""Creates a formatted Excel file from the client data."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_excel(clients: list[dict], company_name: str) -> io.BytesIO:
    """
    Create a polished Excel workbook from the client list.
    Returns a BytesIO buffer ready for download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Client List"

    # ── Styles ──
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2D5F8A")
    data_font = Font(name="Arial", size=10)
    high_fill = PatternFill("solid", fgColor="C6EFCE")
    medium_fill = PatternFill("solid", fgColor="FFEB9C")
    low_fill = PatternFill("solid", fgColor="FFC7CE")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ── Title row ──
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Client Intelligence Report — {company_name}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="2D5F8A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Subtitle row ──
    ws.merge_cells("A2:E2")
    ws["A2"].value = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")

    # ── Summary row ──
    ws.merge_cells("A3:E3")
    high_count = sum(1 for c in clients if c.get("confidence") == "high")
    med_count = sum(1 for c in clients if c.get("confidence") == "medium")
    low_count = sum(1 for c in clients if c.get("confidence") == "low")
    ws["A3"].value = f"Total: {len(clients)} clients  |  High confidence: {high_count}  |  Medium: {med_count}  |  Low: {low_count}"
    ws["A3"].font = Font(name="Arial", size=10, color="555555")
    ws.row_dimensions[3].height = 25

    # ── Headers (row 5) ──
    headers = ["#", "Client Name", "Confidence", "Source", "How Found"]
    col_widths = [6, 35, 14, 20, 50]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[5].height = 28

    # ── Data rows ──
    confidence_fills = {"high": high_fill, "medium": medium_fill, "low": low_fill}

    for row_idx, client in enumerate(clients, start=6):
        row_num = row_idx - 5
        conf = client.get("confidence", "low").lower()
        source_type = client.get("source_type", "website")

        ws.cell(row=row_idx, column=1, value=row_num).font = data_font
        ws.cell(row=row_idx, column=2, value=client.get("name", "")).font = Font(
            name="Arial", size=10, bold=True
        )

        conf_cell = ws.cell(row=row_idx, column=3, value=conf.upper())
        conf_cell.font = Font(name="Arial", size=9, bold=True)
        conf_cell.fill = confidence_fills.get(conf, low_fill)
        conf_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=4, value=source_type).font = data_font
        ws.cell(row=row_idx, column=5, value=client.get("source_context", "")).font = data_font

        # Apply borders
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 5)
            )

    # ── Freeze panes ──
    ws.freeze_panes = "A6"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A5:E{5 + len(clients)}"

    # ── Save to buffer ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
